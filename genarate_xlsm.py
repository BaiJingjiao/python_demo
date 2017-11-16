# -*- coding:utf-8 -*-

from openpyxl import load_workbook
import pysvn
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import shutil 
import time
import re
import fileinput
from os.path import abspath
from inspect import getsourcefile
import json
import pysvn
import tarfile
import os
import shutil
import os.path
import stat
from datetime import datetime
import string
import random

def get_time_stamp(hour, min, sec, formatStr='%Y-%m-%d %X'):
    '''
    %Y-%m-%d %X
    %Y/%m/%d %H:%M:%S.%f
    dateStr = get_time_stamp(0,0,0,'%Y-%m-%d_%H-%M-%S')
    '''
    ISOTIMEFORMAT = formatStr        
    if 0 == hour and 0 == min and 0 == sec:
        stamp = time.strftime(ISOTIMEFORMAT)
    else:
        currentSecond = time.time() 
        newSecond = currentSecond + (hour*3600 + min*60 + sec)  
        stamp = time.strftime(ISOTIMEFORMAT, time.localtime(newSecond)) 
    return stamp

def id_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))

def get_col_num(colName):
    '''
    #通过列名，取得列的索引
    #getColNum('Testcase_Name')
    '''
    #读取模板
    wb = load_workbook('tmss_template.xlsm')
    sheetnames =wb.get_sheet_names() #获得表单名字
    #获得第一个sheet
    sheet = wb.get_sheet_by_name(sheetnames[0])
    for col in range(0,len(sheet.columns)-1):
        if sheet.cell(row=0, column=col).value == colName:
            return col

def get_login(realm, username, may_save):
    return True, 'puccautofactory', 'Huawei123', True

def checkout_cases(svnBranch, casesPath):
    '''
    # casesPath = 'eSDKRest/appSecret/3业务群组管理' 
    '''
    client = pysvn.Client()
    client.callback_get_login = get_login
    svn = r'http://hghsvn01-rd:6801/svn/EBG_UC_Autotest_SVN/UC/Terminal/04 ECS/Project/CloudECS/'
    branchUC = r'CloudECS V600R006/CloudECS V600R006C00_CI/CloudECS V600R006C00_CIB888/Cases/SDV/接口/'
    branchBL = r'CloudECS V600R006/CloudECS V600R006C10/CloudECS V600R006C10B999/Cases/SDV/接口/'
    if 'uc' == svnBranch.lower():
        svnPath = svn + branchUC + casesPath
    elif 'bl' == svnBranch.lower():
        svnPath = svn + branchBL + casesPath
        
    myPath = os.path.dirname(abspath(getsourcefile(lambda:0)))
    dstPath = os.path.join(myPath, 'svn_checkout')
    if os.path.isdir(dstPath):
        shutil.rmtree(dstPath)
    client.checkout(svnPath, dstPath) 
    return dstPath  

def list_all_files(mydir, fileList):
    for root, dirs, files in os.walk(mydir):
        for file in files:
            fileList.append(os.path.join(root, file))
        for d in dirs:
            list_all_files(d, fileList)

def get_tmss_dic(casesPath):
    myPath = os.path.dirname(abspath(getsourcefile(lambda:0)))
    dstPath = os.path.join(myPath, 'svn_checkout')
    fileList = []
#     tmssItems = []
    tmssDic = {}
    list_all_files(dstPath, fileList)
    regex = r'.*\\svn_checkout\\(.*)\\(.*)\\script.py'
    regexItem = r'(.*\\svn_checkout\\)(.*)\\script.py'
    if not casesPath.endswith('/'):
        casesPath = casesPath + '/'
    for item in fileList:
        item = item.decode('cp936').encode('utf8')
        if not(item.endswith('script.py')):
            continue
        n = re.search(regexItem, item)
        if n:
            item = n.groups()[0] + re.sub('##', '', re.sub('/', '\\##', casesPath)) + n.groups()[1] + '\script.py'
#         print 'item', item
        m = re.search(regex, item)
        if m:
            folders = m.groups()[0]
            case = m.groups()[1]
#             #添加'##'以便替换，避免sre_constants.error: bogus escape (end of line) 
#             tmssItems.append(re.sub('##', '', re.sub('/', '\\##', casesPath)) + folders + '\\' + case)
#             key = re.sub('##', '', re.sub('/', '\\##', casesPath)) + folders
#             print 'key', key
            dickey = ''
            for i in folders.split('\\'):
                dickey = dickey + '##' + i
            tmssDic.setdefault(dickey, []).append(case)
#     for key in tmssDic:
#         print key
#         for case in tmssDic[key]:
#             print '-----', case
    return tmssDic

if __name__ == '__main__':
    tmss_version_dic = {
                        'uc':'CloudECS V600R006C00_CIB888',
                        'bl':'CloudECS V600R006C10B999'
                        }
    tmss_topo_dic = {
                     'uc':'【ECS】1ECS_1BMU_SDV_TOPO',
                     'bl':'【ECS】1ECS_1BMU_BASE_SDV_TOPO'
                     }
    print u'\n分支选项:'
    print '-----------------------------'
    for b in tmss_version_dic:
        print b, ':', tmss_version_dic[b]
    print

    branch = raw_input(u'>>>请输入svn分支： '.encode(sys.stdout.encoding))
    branch = branch.strip().lower()
    casesPath = raw_input(u'>>>请输入脚本svn路径： '.encode(sys.stdout.encoding))
    casesPath = casesPath.strip().decode('cp936').encode('utf8')
#     casesPath = 'eSDKRest/10Device管理'
    checkout_cases(branch, casesPath)
    tmssDic = get_tmss_dic(casesPath)
    tmss_Feature_Creation_Version = tmss_version_dic[branch]
    #读取模板
    wb = load_workbook('tmss_template.xlsm', keep_vba=True)
    sheetnames =wb.get_sheet_names() #获得表单名字
    #获得第一个sheet
    sheet = wb.get_sheet_by_name(sheetnames[0])
    tmssColumnDic = {}
    for col in range(len(sheet.columns)):
        tmssColumnDic[(sheet.cell(row=0, column=col).value)] = col
    
    for key in tmssColumnDic:
        print key, '---', tmssColumnDic[key]
    
    therow = 0
    featureSet = {'SDV', '接口'}
    for key in tmssDic:
        tmss_Feature_Name_list = re.sub('^##', '', key).split('##')
        count = 0
        for feature_name in tmss_Feature_Name_list:
            for row in range(len(sheet.rows)):
                featureSet.add(sheet['BD'+str(row+1)].value)
            count += 1
            if feature_name in featureSet:
                continue
            therow += 1
#             print '-'*count+feature_name+'----tmss_Depth: ' + str(count+2)
            sheet.cell(row=therow, column=tmssColumnDic['Depth']).value = '.'*(count+2)
            sheet.cell(row=therow, column=tmssColumnDic['Feature_AnatomyNode']).value = 'FALSE'
            dateStr = get_time_stamp(0,0,0,'%Y-%m-%d %H:%M:%S')
            sheet.cell(row=therow, column=tmssColumnDic['Feature_Creation Version']).value = tmss_Feature_Creation_Version
            if re.search('^\d+', feature_name):
                sheet.cell(row=therow, column=tmssColumnDic['Feature_Name']).value = "'" + feature_name
            else:
                sheet.cell(row=therow, column=tmssColumnDic['Feature_Name']).value = feature_name
            sheet.cell(row=therow, column=tmssColumnDic['Feature_isFeature']).value = 'FALSE'
        for item in tmssDic[key]:
            therow += 1
#             print '----------------'+item + '-----tmss_Depth: ' + str(count+3)
            sheet.cell(row=therow, column=tmssColumnDic['Depth']).value = '.'*(count+3)
            randomStr = get_time_stamp(0,0,0,'%Y%m%d%H%M%S') + id_generator(size=6, chars=string.digits)
            sheet.cell(row=therow, column=tmssColumnDic['Testcase_Automated']).value = 'TRUE'
            sheet.cell(row=therow, column=tmssColumnDic['Testcase_Creation Version']).value = tmss_Feature_Creation_Version
            sheet.cell(row=therow, column=tmssColumnDic['Testcase_Executor Type']).value = 'TEP_xCloud'
            sheet.cell(row=therow, column=tmssColumnDic['Testcase_Level']).value = 'LEVEL1'
            if re.search('^\d+', item):
                sheet.cell(row=therow, column=tmssColumnDic['Testcase_Name']).value = "'" + item
            else:
                sheet.cell(row=therow, column=tmssColumnDic['Testcase_Name']).value = item
            sheet.cell(row=therow, column=tmssColumnDic['Testcase_Number']).value = 'TC.Client.fun.' + randomStr
            sheet.cell(row=therow, column=11).value = tmss_topo_dic[branch] #Testcase_逻辑组网编号
            
    wb.save('your_tmss.xlsm')
    myPath = os.path.dirname(abspath(getsourcefile(lambda:0)))
    print '\n[', os.path.join(myPath, 'your_tmss.xlsm'), ']', u'已生成，请用excel打开并修复，再导入TMSS!'
